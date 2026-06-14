import openpyxl
import json
import requests
import datetime
import os
import subprocess
import re

# Configuración del Google Sheet
SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/17r2QvAAQsZZK7YB_UXstmDObZC7WkofGII_cbLUVybU/export?format=xlsx"
LOCAL_XLSX_PATH = "google_sheet_temp.xlsx"
DATA_JS_PATH = "data.js"

# Mapeo de días de la semana
DIA_MAP = {
    "lunes": "lun",
    "martes": "mar",
    "miercoles": "mie",
    "miércoles": "mie",
    "jueves": "jue",
    "viernes": "vie",
    "sabado": "sab",
    "sábado": "sab",
    "domingo": "dom"
}

# Mapeo de meses según pestañas
MES_MAP = {
    "Mayo25": "may",
    "JUN25": "jun",
    "JUL25": "jul",
    "Agos25": "ago",
    "SEP25": "sep",
    "Oct25": "oct",
    "Nov25": "nov",
    "Dic25": "dic",
    "Ene26": "ene",
    "Feb26": "feb",
    "Marz26": "mar",
    "Abril26": "abr",
    "May26": "may",
    "Jun26": "jun"
}

# El orden en el que se listan los meses en el dashboard comparativo
MONTH_ORDER = [
    "Mayo25", "JUN25", "JUL25", "Agos25", "SEP25", "Oct25", "Nov25", "Dic25",
    "Ene26", "Feb26", "Marz26", "Abril26", "May26", "Jun26"
]

# Datos históricos de Outlet extraídos de index.html original
OUTLET_HISTORIC_DATA = [
    # Datos de Febrero 2026
    ["2 feb", None, 39200, -39200],
    ["3 feb", 12000, 35000, -23000],
    ["4 feb", 50580, 42900, 7680],
    ["5 feb", 40000, 40220.68, -220.68],
    ["6 feb", 76900, 35150, 41750],
    ["7 feb", 183000, 40000, 143000],
    ["9 feb", 504698, 35400, 469298],
    ["10 feb", 26750, 105000, -78250],
    ["11 feb", 67500, 35000, 32500],
    ["12 feb", 15000, 1500, 13500],
    ["13 feb", 69000, 44000, 25000],
    ["14 feb", 79200, 0, 79200],
    ["17 feb", 15000, 25000, -10000],
    ["18 feb", 68000, 32000, 36000],
    ["19 feb", 0, 65000, -50000],
    ["20 feb", 15000, 70000, -55000],
    ["21 feb", 97900, 65000, 32900],
    ["23 feb", 40000, 35000, 5000],
    ["24 feb", 91500, 32600, 58900],
    ["25 feb", 184000, 1500, 182500],
    ["26 feb", 186000, 49500, 136500],
    ["27 feb", 171300, 35000, 136300],
    ["28 feb", 118000, 0, 118000],
    # Datos de Marzo 2026
    ["2 mar", 30000, 55000, -25000],
    ["3 mar", 50910, 70400, -19490],
    ["4 mar", 0, 55000, -55000],
    ["5 mar", 30000, 64500, -34500],
    ["6 mar", 0, 60000, -60000],
    ["7 mar", 183000, 40000, 143000],
    ["9 mar", 247500, 60400, 187100],
    ["10 mar", 38000, 96300, -58300],
    ["11 mar", 160000, 63000, 97000],
    ["12 mar", 28000, 58800, -30800],
    ["13 mar", 41100, 66500, -25400],
    ["14 mar", 183000, 0, 183000],
    ["16 mar", 10000, 55000, -45000],
    ["17 mar", 64000, 67500, -3500],
    ["18 mar", 49200, 55000, -5800],
    ["19 mar", 0, 55000, -55000],
    ["21 mar", 0, 40000, -40000],
    ["23 mar", 36000, 65000, -29000],
    ["24 mar", 0, 60000, -60000],
    ["25 mar", 0, 55000, -55000],
    ["26 mar", 64300, 35000, 29300],
    ["27 mar", 44100, 52990, -8890],
    ["28 mar", 0, 40800, -40800],
    ["30 mar", 43000, 55000, -12000],
    ["31 mar", 156200, 56200, 100000]
]

def clean_num(val):
    if val is None or str(val).strip() == "" or str(val).lower() == "nan" or str(val).lower() == "null":
        return 0
    try:
        s = str(val).strip().replace("$", "").replace(" ", "")
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return 0

def is_valid_day_num(val):
    if val is None:
        return False
    try:
        f_val = float(str(val).strip())
        return f_val.is_integer() and 1 <= f_val <= 31
    except ValueError:
        return False

def download_sheet():
    print("Descargando Google Sheet en formato XLSX...")
    r = requests.get(SPREADSHEET_URL)
    if r.status_code == 200:
        with open(LOCAL_XLSX_PATH, "wb") as f:
            f.write(r.content)
        print("Descarga completada.")
        return True
    else:
        print(f"Error al descargar: {r.status_code}")
        return False

def parse_month_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f"Advertencia: Pestaña '{sheet_name}' no encontrada en el Excel.")
        return []
    
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    
    daily_records = []
    
    current_day_num = None
    current_day_name = ""
    current_turno = "" # "MAÑANA" o "TARDE"
    
    day_data = None
    
    for idx, row in enumerate(rows):
        if not row:
            continue
            
        col_a = str(row[0]).strip().lower() if row[0] is not None else ""
        col_b = row[1]
        col_c = str(row[2]).strip().lower() if row[2] is not None else ""
        col_d = row[3]
        
        # 1. Detectar cambio de día
        if col_a in DIA_MAP and is_valid_day_num(col_b):
            if day_data and day_data["day_num"] is not None:
                daily_records.append(day_data)
                
            current_day_num = int(float(str(col_b).strip()))
            current_day_name = DIA_MAP[col_a]
            current_turno = ""
            
            day_data = {
                "day_num": current_day_num,
                "day_name": current_day_name,
                "month_code": MES_MAP[sheet_name],
                "ventas_tm": 0,
                "gastos_tm": 0,
                "empleado_tm": None,
                "ventas_tt": 0,
                "gastos_tt": 0,
                "empleado_tt": None
            }
            continue
            
        if day_data is None:
            continue
            
        # 2. Detectar Turno
        if "turno maña" in col_a or "primera maña" in col_a:
            current_turno = "MAÑANA"
            continue
        elif "turno tarde" in col_a:
            current_turno = "TARDE"
            continue
            
        # 3. Detectar Cierres / Totales de Ventas y Gastos
        if current_turno == "MAÑANA":
            if any(x in col_c or x in col_a for x in ["venta total tm", "facturacion total tm", "facturación total tm"]):
                day_data["ventas_tm"] = clean_num(col_d)
            elif any(x in col_c or x in col_a for x in ["gastos total tm", "gasto total tm"]):
                day_data["gastos_tm"] = clean_num(col_d)
            elif "sobrente de caja" in col_c or "sobrantede caja" in col_c or "sobrante/faltante" in col_c:
                emp = str(row[0]).strip() if row[0] is not None else ""
                if emp and emp.lower() not in ["", "nan", "null"]:
                    day_data["empleado_tm"] = emp.upper()
                    
        elif current_turno == "TARDE":
            if any(x in col_c or x in col_a for x in ["venta total tt", "facturacion total tt", "facturación total tt", "venta total tm", "facturacion total tm"]):
                day_data["ventas_tt"] = clean_num(col_d)
            elif any(x in col_c or x in col_a for x in ["gastos total tm", "gastos total tt", "gasto total tm", "gastos total del dia", "gastos total del día", "gasto total del dia"]):
                day_data["gastos_tt"] = clean_num(col_d)
            elif "sobrante/faltante" in col_c or "sobrantede caja" in col_c:
                emp = str(row[0]).strip() if row[0] is not None else ""
                if emp and emp.lower() not in ["", "nan", "null"]:
                    day_data["empleado_tt"] = emp.upper()

    if day_data and day_data["day_num"] is not None:
        daily_records.append(day_data)
        
    asistencia = {}
    for r in daily_records:
        asistencia[r["day_num"]] = {
            "empleado_tm": r["empleado_tm"] or "—",
            "empleado_tt": r["empleado_tt"] or "—",
            "ventas_tm": r["ventas_tm"],
            "ventas_tt": r["ventas_tt"]
        }
        
    return asistencia

MONTH_YEAR_MAP = {
    "Mayo25": (2025, 5),
    "JUN25": (2025, 6),
    "JUL25": (2025, 7),
    "Agos25": (2025, 8),
    "SEP25": (2025, 9),
    "Oct25": (2025, 10),
    "Nov25": (2025, 11),
    "Dic25": (2025, 12),
    "Ene26": (2026, 1),
    "Feb26": (2026, 2),
    "Marz26": (2026, 3),
    "Abril26": (2026, 4),
    "May26": (2026, 5),
    "Jun26": (2026, 6)
}

# Correcciones específicas para montos con ceros de más (typos en planillas mensuales)
FINANCIAL_CORRECTIONS = {
    (2025, 6, 21): {
        "ventas_tt": 52447.95,
        "gastos_tt": 92350.0
    },
    (2025, 7, 3): {
        "ventas_tm": 44979.59
    },
    (2025, 7, 8): {
        "ventas_tm": 266737.60
    },
    (2025, 7, 17): {
        "ventas_tm": 36275.18
    },
    (2025, 7, 18): {
        "ventas_tm": 81775.22
    },
    (2025, 7, 26): {
        "ventas_tm": 21280.00,
        "ventas_tt": 30907.47
    }
}

def parse_cierre_auditoria(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f"Advertencia: Hoja '{sheet_name}' no encontrada para auditoría.")
        return []
        
    ws = wb[sheet_name]
    auditoria = []
    
    current_day_num = None
    current_day_name = ""
    
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
            
        col_a = row[0]
        col_b = row[1]
        col_c = row[2]
        col_d = row[3]
        
        col_a_str = str(col_a).strip().lower() if col_a is not None else ""
        col_b_str = str(col_b).strip() if col_b is not None else ""
        col_c_str = str(col_c).strip() if col_c is not None else ""
        
        # Detectar día para poder etiquetar la fila
        if col_a_str in DIA_MAP and is_valid_day_num(col_b):
            current_day_num = int(float(col_b_str))
            current_day_name = DIA_MAP[col_a_str]
            
        # Filtrar exclusivamente las filas que digan "VENTA TOTAL TM" o "VENTA TOTAL TT" (ignorando espacios iniciales/finales)
        if "venta total tm" in col_c_str.lower() or "venta total tt" in col_c_str.lower():
            monto = clean_num(col_d)
            day_label = f"{current_day_name} {current_day_num}" if current_day_num else "—"
            # Guardamos: [nro_fila, day_label, concepto (original), monto]
            auditoria.append([
                idx,
                day_label,
                str(col_c),
                monto
            ])
            
    print(f"Total de filas de auditoría extraídas de {sheet_name}: {len(auditoria)}")
    return auditoria

def parse_employee_expenses(wb, sheet_name):
    # Encontrar la pestaña de manera flexible
    sheet = None
    for name in wb.sheetnames:
        if name.lower().replace('/', '') == sheet_name.lower().replace('/', ''):
            sheet = wb[name]
            break
    if not sheet:
        return []
        
    # 1. Obtener asistencia de personal de los cierres (columna A de Sobrante/Faltante)
    asistencia = {} # { (day_num, turno): employee_name }
    
    current_day_num = None
    current_day_name = ""
    current_turno = ""
    
    for r_idx in range(1, sheet.max_row + 1):
        colA = sheet.cell(row=r_idx, column=1).value
        colB = sheet.cell(row=r_idx, column=2).value
        colC = sheet.cell(row=r_idx, column=3).value
        
        colA_str = str(colA).strip().lower() if colA is not None else ""
        colC_str = str(colC).strip().lower() if colC is not None else ""
        
        is_novedades = "novedades del dia" in colC_str or "novedades" in colC_str
        is_new_day = (colA_str in DIA_MAP and is_valid_day_num(colB)) or (is_novedades and is_valid_day_num(colB))
        
        if is_new_day:
            try:
                current_day_num = int(float(str(colB).strip()))
            except:
                current_day_num = None
            current_turno = ""
            continue
            
        if "turno maña" in colA_str or "primera maña" in colA_str:
            current_turno = "MAÑANA"
            continue
        elif "turno tarde" in colA_str:
            current_turno = "TARDE"
            continue
            
        if current_day_num and current_turno:
            is_sobrante = "sobrante/faltante" in colC_str or "sobrantede caja" in colC_str or "sobrente de caja" in colC_str
            if is_sobrante and colA:
                emp = str(colA).strip().upper()
                if emp and emp.lower() not in ["", "nan", "null"]:
                    asistencia[(current_day_num, current_turno)] = emp

    # 2. Buscar las filas de total y aplicar la regla
    current_day_num = None
    current_day_name = ""
    current_turno = ""
    expenses = []
    
    for r_idx in range(1, sheet.max_row + 1):
        colA = sheet.cell(row=r_idx, column=1).value
        colB = sheet.cell(row=r_idx, column=2).value
        colC = sheet.cell(row=r_idx, column=3).value
        
        colA_str = str(colA).strip().lower() if colA is not None else ""
        colC_str = str(colC).strip().lower() if colC is not None else ""
        
        is_novedades = "novedades del dia" in colC_str or "novedades" in colC_str
        is_new_day = (colA_str in DIA_MAP and is_valid_day_num(colB)) or (is_novedades and is_valid_day_num(colB))
        
        if is_new_day:
            try:
                current_day_num = int(float(str(colB).strip()))
            except:
                current_day_num = None
            if colA_str in DIA_MAP:
                current_day_name = DIA_MAP[colA_str]
            else:
                current_day_name = "—"
            current_turno = ""
            continue
            
        if "turno maña" in colA_str or "primera maña" in colA_str:
            current_turno = "MAÑANA"
            continue
        elif "turno tarde" in colA_str:
            current_turno = "TARDE"
            continue
            
        if not current_day_num or not current_turno:
            continue
            
        # Detectar filas de total de ventas o gastos de turno
        is_total_tm = "venta total tm" in colC_str or "facturacion total tm" in colC_str or "facturación total tm" in colC_str or "venta total tm" in colA_str or "facturacion total tm" in colA_str
        is_total_tt = "venta total tt" in colC_str or "facturacion total tt" in colC_str or "facturación total tt" in colC_str or "venta total tt" in colA_str or "facturacion total tt" in colA_str or "gasto total del dia" in colC_str or "gasto total del dia" in colA_str
        
        is_gasto_tm = "gasto total tm" in colC_str or "gasto total tm" in colA_str or "gasto total del dia" in colC_str or "gasto total del dia" in colA_str or "gastos total tm" in colC_str or "gastos total tm" in colA_str
        is_gasto_tt = "gasto total tt" in colC_str or "gasto total tt" in colA_str or "gastos total tt" in colC_str or "gastos total tt" in colA_str
        
        is_cierre_row = is_total_tm or is_total_tt or is_gasto_tm or is_gasto_tt
        
        if is_cierre_row:
            valJ = sheet.cell(row=r_idx, column=10).value
            valK = sheet.cell(row=r_idx, column=11).value
            valL = sheet.cell(row=r_idx, column=12).value
            
            gastoJ = clean_num(valJ)
            gastoK = clean_num(valK)
            total = gastoJ + gastoK
            
            if total > 0:
                emp_name = "—"
                # Regla: si en col L hay un nombre importarlo, sino sacar de asistencia del turno
                if valL is not None and not isinstance(valL, (int, float)) and str(valL).strip():
                    valL_str = str(valL).strip()
                    valL_lower = valL_str.lower()
                    is_money = "$" in valL_str or (valL_lower.replace('-','').replace('.','').replace(',','').isdigit() and len(valL_str) > 0)
                    is_stock_header = any(x in valL_lower for x in ["stock", "showroom", "outlet"])
                    is_omit = valL_lower in ["", "empleado", "nombre", "nan", "null"]
                    
                    if not is_money and not is_stock_header and not is_omit:
                        emp_name = valL_str.upper()
                    else:
                        emp_name = asistencia.get((current_day_num, current_turno), "—")
                else:
                    emp_name = asistencia.get((current_day_num, current_turno), "—")
                
                day_label = f"{current_day_name} {current_day_num}"
                expenses.append([
                    sheet_name,
                    r_idx,
                    day_label,
                    current_turno,
                    str(colC or colA).strip(),
                    emp_name,
                    gastoJ,
                    gastoK,
                    total
                ])
                
    return expenses

def compile_showroom_data_from_months(wb):
    showroom_data = []
    
    DIA_MAP_ES = {0: "lun", 1: "mar", 2: "mie", 3: "jue", 4: "vie", 5: "sab", 6: "dom"}
    MES_MAP_ES = {
        1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
        7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic"
    }
    
    for month_sheet in MONTH_ORDER:
        if month_sheet not in wb.sheetnames:
            continue
        sheet = wb[month_sheet]
        rows = list(sheet.iter_rows(values_only=True))
        
        daily_records = []
        current_day_num = None
        current_day_name = ""
        current_turno = ""
        day_data = None
        
        for idx, row in enumerate(rows):
            if not row:
                continue
                
            col_a = str(row[0]).strip().lower() if row[0] is not None else ""
            col_b = row[1]
            col_c = str(row[2]).strip().lower() if row[2] is not None else ""
            col_d = row[3]
            
            # Detectar cambio de día
            is_novedades = "novedades del dia" in col_c or "novedades" in col_c
            is_new_day = (col_a in DIA_MAP and is_valid_day_num(col_b)) or (is_novedades and is_valid_day_num(col_b))
            
            if is_new_day:
                if day_data and day_data["day_num"] is not None:
                    daily_records.append(day_data)
                    
                try:
                    current_day_num = int(float(str(col_b).strip()))
                except:
                    current_day_num = None
                current_day_name = DIA_MAP.get(col_a, "—")
                current_turno = ""
                
                day_data = {
                    "day_num": current_day_num,
                    "day_name": current_day_name,
                    "ventas_tm": 0,
                    "gastos_tm": 0,
                    "empleado_tm": None,
                    "ventas_tt": 0,
                    "gastos_tt": 0,
                    "empleado_tt": None
                }
                continue
                
            if day_data is None:
                continue
                
            # Detectar Turno
            if "turno maña" in col_a or "primera maña" in col_a:
                current_turno = "MA\u00d1ANA"
                continue
            elif "turno tarde" in col_a:
                current_turno = "TARDE"
                continue
                
            # Detectar Cierres / Totales de Ventas y Gastos
            if current_turno == "MA\u00d1ANA":
                if any(x in col_c or x in col_a for x in ["venta total tm", "facturacion total tm", "facturación total tm"]):
                    day_data["ventas_tm"] = clean_num(col_d)
                elif any(x in col_c or x in col_a for x in ["gastos total tm", "gasto total tm"]):
                    day_data["gastos_tm"] = clean_num(col_d)
                elif "sobrente de caja" in col_c or "sobrantede caja" in col_c or "sobrante/faltante" in col_c:
                    emp = str(row[0]).strip() if row[0] is not None else ""
                    if emp and emp.lower() not in ["", "nan", "null"]:
                        day_data["empleado_tm"] = emp.upper()
                        
            elif current_turno == "TARDE":
                if any(x in col_c or x in col_a for x in ["venta total tt", "facturacion total tt", "facturación total tt", "venta total tm", "facturacion total tm"]):
                    day_data["ventas_tt"] = clean_num(col_d)
                elif any(x in col_c or x in col_a for x in ["gastos total tm", "gastos total tt", "gasto total tm", "gastos total del dia", "gastos total del día", "gasto total del dia"]):
                    day_data["gastos_tt"] = clean_num(col_d)
                elif "sobrante/faltante" in col_c or "sobrantede caja" in col_c:
                    emp = str(row[0]).strip() if row[0] is not None else ""
                    if emp and emp.lower() not in ["", "nan", "null"]:
                        day_data["empleado_tt"] = emp.upper()
                        
        if day_data and day_data["day_num"] is not None:
            daily_records.append(day_data)
            
        # Convertir daily_records a registros de showroom_data
        y, m = MONTH_YEAR_MAP[month_sheet]
        for r in daily_records:
            d = r["day_num"]
            
            # Aplicar correcciones específicas si existen para esta fecha
            fecha_key = (y, m, d)
            if fecha_key in FINANCIAL_CORRECTIONS:
                for campo, valor in FINANCIAL_CORRECTIONS[fecha_key].items():
                    r[campo] = valor
            
            ventas = r["ventas_tm"] + r["ventas_tt"]
            gastos = r["gastos_tm"] + r["gastos_tt"]
            ganancias = ventas - gastos
            
            # Omitir si ventas y gastos son cero
            if ventas == 0 and gastos == 0:
                continue
                
            emp_tm = r["empleado_tm"] or "\u2014"
            emp_tt = r["empleado_tt"] or "\u2014"
            
            # Formatear el label con el año
            try:
                date_obj = datetime.date(y, m, d)
                day_name = DIA_MAP_ES[date_obj.weekday()]
            except Exception as e:
                day_name = r["day_name"] if r["day_name"] != "\u2014" else "\u2014"
                
            month_code = MES_MAP_ES[m]
            year_code = str(y)[-2:]
            label = f"{day_name} {d} {month_code} {year_code}"
            
            showroom_data.append([
                label,
                ventas,
                gastos,
                ganancias,
                emp_tm,
                emp_tt
            ])
            
    return showroom_data

def main():

    if not download_sheet():
        print("No se pudo obtener el archivo de Google Sheets. Cancelando sincronización.")
        return
        
    wb = openpyxl.load_workbook(LOCAL_XLSX_PATH, data_only=True)
    
    # 1. Obtener la asistencia de personal de todas las pestañas mensuales
    asistencia_por_mes = {}
    print("Procesando asistencia de personal en pestañas mensuales...")
    for month_sheet in MONTH_ORDER:
        print(f"Procesando asistencia de {month_sheet}...")
        asistencia = parse_month_sheet(wb, month_sheet)
        if month_sheet in MONTH_YEAR_MAP:
            key = MONTH_YEAR_MAP[month_sheet]
            asistencia_por_mes[key] = asistencia
            
    # 2. Compilar los datos financieros oficiales de Showroom directamente desde las pestañas mensuales (Balance)
    print("Compilando datos financieros de Showroom desde las pestañas mensuales...")
    showroom_data = compile_showroom_data_from_months(wb)
    print(f"Total de registros de Showroom consolidados desde Balance: {len(showroom_data)}")
    
    # 3. Generar la solapa exclusiva de auditoría de turnos para Junio, Mayo, Abril y Marzo 2026
    cierre_junio_data = parse_cierre_auditoria(wb, "Jun26")
    cierre_mayo_data = parse_cierre_auditoria(wb, "May26")
    cierre_abril_data = parse_cierre_auditoria(wb, "Abril26")
    cierre_marzo_data = parse_cierre_auditoria(wb, "Marz26")
    
    # 4. Obtener todos los gastos/adelantos de empleados de todos los meses de la temporada
    gastos_empleados_data = []
    print("Procesando adelantos y gastos de empleados en pestañas mensuales...")
    for month_sheet in MONTH_ORDER:
        print(f"Procesando gastos de {month_sheet}...")
        expenses = parse_employee_expenses(wb, month_sheet)
        gastos_empleados_data.extend(expenses)
    print(f"Total de registros de adelantos extraídos: {len(gastos_empleados_data)}")
            
    # Generar el archivo data.js
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    js_content = f"""// Archivo generado automáticamente por el script de sincronización.
// Última actualización: {timestamp}

const DATOS_SINCRONIZADOS = {{
  ultima_actualizacion: "{timestamp}",
  showroom: {json.dumps(showroom_data)},
  outlet: {json.dumps(OUTLET_HISTORIC_DATA)},
  cierre_junio_2026: {json.dumps(cierre_junio_data)},
  cierre_mayo_2026: {json.dumps(cierre_mayo_data)},
  cierre_abril_2026: {json.dumps(cierre_abril_data)},
  cierre_marzo_2026: {json.dumps(cierre_marzo_data)},
  gastos_empleados: {json.dumps(gastos_empleados_data)}
}};
"""
    
    with open(DATA_JS_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)
        
    print(f"Archivo '{DATA_JS_PATH}' escrito con éxito.")
    
    if os.path.exists(LOCAL_XLSX_PATH):
        os.remove(LOCAL_XLSX_PATH)
        
    try:
        print("Comprobando cambios en Git...")
        status = subprocess.check_output(["git", "status", "--porcelain"], text=True)
        if "data.js" in status or "package.json" in status or "sync_sheets.py" in status:
            print("Subiendo cambios actualizados a GitHub...")
            subprocess.run(["git", "add", "data.js", "package.json", "sync_sheets.py"])
            subprocess.run(["git", "commit", "-m", f"Sincronización automática de Google Sheets - {timestamp}"])
            subprocess.run(["git", "push", "origin", "main"])
            print("Push a GitHub completado con éxito.")
        else:
            print("No hay cambios en los datos. No se requiere hacer push.")
    except Exception as e:
        print(f"Error al subir cambios a GitHub: {str(e)}")

if __name__ == "__main__":
    main()
